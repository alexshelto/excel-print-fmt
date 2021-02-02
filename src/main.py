import openpyxl
from pathlib import Path
import argparse
from typing import Optional, Sequence



def generate(sheet, exclude):
    # Generating column labels for output
    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)
    
    # Displaying the columns that are being excluded
    print(f'excluding columns: {[col_names[i] for i in range(0,len(col_names)) if i in exclude]}', end='\n\n')

    # formatted output logic
    row_num = 0
    for row in sheet.iter_rows():
        cell_num = 0
        if row_num != 0:
            for cell in row:
                if cell_num not in exclude and cell.value is not None:
                    print(f'{col_names[cell_num]}: {cell.value}', end='\n')
                cell_num += 1
            print('#' * 20)
        row_num += 1



def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('filename', help='filename or relative path')
    parser.add_argument('-e','--exclude', nargs='+', help='columns to ignore')

    args = parser.parse_args()
    exclude_nums = [int(i) for i in args.exclude] if args.exclude != None else []

    #xlsx_file = Path('../test', 'test1.xlsx')
    try:
        wb_obj = openpyxl.load_workbook(args.filename)
        sheet = wb_obj.active
    except:
        print('Error loading file, check your path to the file')
        return 1
    
    generate(sheet, set(exclude_nums))
    return 0

if __name__ == '__main__':
    exit(main())
